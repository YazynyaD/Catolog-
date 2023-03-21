import array as arr
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import importlib.util
import sys
import eel
import os
import pandas as pd
from openpyxl import *
# Планы: 
# 1. Другая вкладка с выбором каталогов
# 2. Прием письма с сайта
# 3. Список автокаталогов без капчи


z=0
ci=0
service = Service(executable_path="/geckodriver")
cii=ci

eel.init("web")
@eel.expose
def beackup():
    try:
        wb = load_workbook('./beackdata.xlsx')
        sheet = wb['Sheet1']
        namecl11 = sheet['B2'].value
        # print(namecl11)
        city11= sheet['B3'].value
        cityr11= sheet['B4'].value
        ciadress11= sheet['B5'].value
        street11= sheet['B6'].value
        home11= sheet['B7'].value
        adress11= sheet['B8'].value
        numclinic11= sheet['B9'].value
        ind11= sheet['B10'].value
        fio11= sheet['B11'].value
        timework11= sheet['B12'].value
        url11= sheet['B13'].value
        mail11= sheet['B14'].value
        keysl11= sheet['B15'].value
        passw11= sheet['B16'].value
        desc11= sheet['B17'].value
        tupecl11= sheet['B18'].value
        eel.beackupsh(namecl11, city11, cityr11,ciadress11,street11, home11, adress11,numclinic11,ind11,fio11,timework11,url11,mail11,keysl11,passw11,desc11,tupecl11)
        return 0
    
    except Exception:
        print("Ошибка выгрузки")
        pass
@eel.expose
def conti(status,ur,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind,sites):
    

    df = pd.DataFrame({'Данные': [namecl, city, cityr,ciadress,street, home, adress,numclinic,ind,fio,timework,url,mail,keysl,passw,desc,tupecl]})
    df.to_excel('./beackdata.xlsx')
    
    wb = load_workbook('./express-result.xlsx')

    sheet = wb['Sheet1']
    list_B2 = sheet['B1'].value
    print(list_B2)
    print(sites)
    numclinic="8"+numclinic
    numpl="+7"+numpl
    numcodcity=numcodcity[:-7]
    numclshot=numclshot[3:]
    numtrue=numn
    if fio.count(" ") == 2:
        fiom=fio.split(" ")
        f=fiom[0]
        i=fiom[1]
        o=fiom[2]
    else:
        f="Дронов"
        i="Михаил"
        o="Иванович"
    unamecl="ООО"+" \""+namecl+"\""
    print(unamecl)
    namepers=i+" "+f
    if mail.count("@") == 1:
        uurl=url[:-4]
        uurl=uurl[8:]
        uurl = uurl.replace(".", "-")
        uurl = uurl.replace("_", "-")
        mmail=mail.split("@")
        login="login-"+mmail[0]+"123"
        print("Логин(Ник): "+login)
        # print(uurl)
    if namecl !="":
        print("Название: "+namecl)
    else:
        print("Название: !!!")
    if url !="":
        print("Ссылка на сайт: "+url)
    else:
        print("Ссылка на сайт: !!!")
    if len(numclinic)>1:
        print("Номер: "+numclinic[:-10]+"("+numclinic[1:-7]+")"+numclinic[4:-4]+"-"+numclinic[7:-2]+"-"+numclinic[9:])
    else:
        print("Номер: !!!")
    if city !="":
        print("Город, область: "+city+", "+cityr)
    else:
        print("Город, область: !!!")
    if ciadress !="":
        print("Полный адрес: "+ciadress)
    else:
        print("Полный адрес: !!!")
    if adress !="":
        print("Адрес: "+adress)
    else:
        print("Адрес: !!!")
    if street !="":
        print("Улица: "+street)
    else:
        print("Улица: !!!")
    if home !="":
        print("Дом: "+home)
    else:
        print("Дом: !!!")
    print("Пароль: "+passw)
    print("ФИО: "+fio)
    print("Сфера: "+tupecl)
    if tupecl == "":
        tupecl="Клиника наркологической помощи"
    if keysl == "":
        keysl="нарколог на дом, вывод из запоя, лечение от алкоголизма, кодирование от алкоголизма, наркологическая помощь, лечение наркомании"
    if timework == "":
        timework="Круглосуточно, ежедневно"
    if passw == "":
        passw="LW9Tx44d"
    if ind=="":
        service = Service(executable_path="/geckodriver")
        # os.environ['MOZ_HEADLESS']='1'
        # options.headless = True
        options = Options()
        options.add_argument("--headless")
        brow1=webdriver.Firefox(service=service,options=options)
        brow1.get(url="https://www.pochta.ru/RU/post-index")
        sleep(2)
        try:
            zz=brow1.find_element(By.XPATH, "/html/body/div[6]/div/div/div[2]/div/div/div/div/div[1]/div/div/div/input")
            zz.click()
            zz.clear()
            zz.send_keys(ciadress)
            sleep(1)
            zz.send_keys(Keys.DOWN)
            sleep(1)
            zz.send_keys(Keys.ENTER)
            sleep(2)
            ind=brow1.find_element(By.CLASS_NAME, "index-card__postal-code").text
        except Exception:
            # print("Ошибка: найти индекс")
            pass
        brow1.quit()
        
        # print("Индекс: "+ind+"\n\n")
        if ind !="":
            print("Индекс: "+ind+"\n\n")
        else:
            print("Индекс: !!!\n")

    # print(tupecl)
    if "\n" in status:
        statuss=status.split("\n")
        for stat in statuss:
            print(stat[:-1])
            popit(stat,ur,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind,sites)
    else:
        popit(status,ur,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind,sites)

@eel.expose
def opsite():
    eel.show("site.html")
    
def popit(izzi,ur,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind,sites):
    print("Процесс...")
    izzi=izzi+"\n"
    urr = os.path.abspath('')
    urr=urr.replace("\\", "/")
    urr=str(urr)
    ur=urr
    print(izzi)
    if cii==ci:
        if izzi=="Все\n" or izzi=="все\n" or izzi=="all\n" or izzi=="All\n":
            print("Вызов \"Все\"\n\n")
            zz=1
            service = Service(executable_path="/geckodriver")
            # os.environ['MOZ_HEADLESS']='False'
            brow=webdriver.Firefox(service=service)
            for iz in range(1,140):
                if iz==19 or iz==38 or iz==57 or iz==76 or iz==95 or iz==114 or iz==133:
                    service = Service(executable_path="/geckodriver")
                    # os.environ['MOZ_HEADLESS']='False'
                    # options.headless = False
                    brow=webdriver.Firefox(service=service)
                    zz=1
                spec = importlib.util.spec_from_file_location(f"module.q{iz}q", f"{ur}/st/q{iz}q.py")
                foo1 = importlib.util.module_from_spec(spec)
                sys.modules[f"module.q{iz}q"] = foo1
                spec.loader.exec_module(foo1)
                foo1.goo(brow,zz-1,ur,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind)
                zz=zz+1
                
        elif izzi=="Отрезок\n" or izzi=="Интервал\n" or izzi=="интервал\n" or izzi=="отрезок\n" or izzi=="отр\n":
            otr1=int(otr1)
            otr2=int(otr2)
            print("Вызов \"Отрезок с "+str(otr1)+" по "+str(otr2)+"\"\n\n")
            service = Service(executable_path="/geckodriver")
            brow=webdriver.Firefox(service=service)
            izi=zz=1
            for iz in range(otr1,otr2+1):
                if iz==19 or iz==38 or iz==57 or iz==76 or iz==95 or iz==114 or iz==133:
                    service = Service(executable_path="/geckodriver")
                    brow=webdriver.Firefox(service=service)
                    zz=1
                izi=izi+1
                spec = importlib.util.spec_from_file_location(f"module.q{iz}q", f"{urr}/st/q{iz}q.py")
                foo1 = importlib.util.module_from_spec(spec)
                sys.modules[f"module.q{iz}q"] = foo1
                spec.loader.exec_module(foo1)
                foo1.goo(brow,zz-1,urr,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind)
                zz=zz+1
                
        elif izzi=="Кол-во\n" or izzi=="кол-во\n" or izzi=="количество\n" or izzi=="Количество\n":
            # print("if(Кол-во)")
            otr1=int(otr1)
            print("Вызов \"Количество "+str(otr1)+" шт.""\"\n\n")
            zz=1
            service = Service(executable_path="/geckodriver")
            brow=webdriver.Firefox(service=service)
            for iz in range(1,otr1+1):
                if iz==19 or iz==38 or iz==57 or iz==76 or iz==95 or iz==114 or iz==133:
                    service = Service(executable_path="/geckodriver")
                    brow=webdriver.Firefox(service=service)
                    zz=1
                spec = importlib.util.spec_from_file_location(f"module.q{iz}q", f"{urr}/st/q{iz}q.py")
                foo1 = importlib.util.module_from_spec(spec)
                sys.modules[f"module.q{iz}q"] = foo1
                spec.loader.exec_module(foo1)
                foo1.goo(brow,zz,urr,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind)
                zz=zz+1
                
        elif izzi=="Определенные сайты\n" or izzi=="Определенные\n" or izzi=="Определенные сайты":
            print("Вызов \"Определенные сайты\"\n\n")
            service = Service(executable_path="/geckodriver")
            brow=webdriver.Firefox(service=service)
            izi=zz=1
            for iz in range(1,len(sites)+1):
                if iz==19 or iz==38 or iz==57 or iz==76 or iz==95 or iz==114 or iz==133:
                    service = Service(executable_path="/geckodriver")
                    brow=webdriver.Firefox(service=service)
                    zz=1
                izi=izi+1
                spec = importlib.util.spec_from_file_location(f"module.q{sites[zz-1]}q", f"{urr}/st/q{sites[zz-1]}q.py")
                foo1 = importlib.util.module_from_spec(spec)
                sys.modules[f"module.q{sites[zz-1]}q"] = foo1
                spec.loader.exec_module(foo1)
                foo1.goo(brow,zz-1,urr,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind)
                zz=zz+1
        elif izzi=="Экспресс(beta)\n" or izzi=="Экспрес\n":
            sss=[2,4,5,6,9,11,13,16,17,19,21,24,26,28,30,32,35,41,42]
            # 2,4,5,6,9,
            print("Вызов \"Экспресс+выгрузка\"\n\n")
            service = Service(executable_path="/geckodriver")
            brow=webdriver.Firefox(service=service)
            izi=zz=1
            b2=111
            ar=[0, 1, 2, 3,4, 5,6,7,8,9,10,11,12,13,14,15,16,17,18]
            for iz in range(1,len(sss)+1):
                if iz==19 or iz==38 or iz==57 or iz==76 or iz==95 or iz==114 or iz==133:
                    service = Service(executable_path="/geckodriver")
                    brow=webdriver.Firefox(service=service)
                    zz=1
                izi=izi+1
                spec = importlib.util.spec_from_file_location(f"module.q{sss[iz-1]}q", f"{urr}/st/autost/q{sss[iz-1]}q.py")
                foo1 = importlib.util.module_from_spec(spec)
                sys.modules[f"module.q{sss[iz-1]}q"] = foo1
                spec.loader.exec_module(foo1)
                b2="Nne"
                foo1.goo(brow,zz-1,urr,otr1,otr2,urli,notwww,namecl,unamecl,city,cityr,ciadress,street,home,adress,numclinic,numtrue,namepers,fio,f,i,o,timework,url,mail,keysl,desc,tupecl,passw,login,numcodcity,numclshot,numn,numpl,shcir,regi,ob,bb,ind)
                wb = load_workbook('./st/autost/ex/ur.xlsx')
                sheet = wb['Sheet1']
                b2 = sheet['B2'].value
            
                # print(str(ar[iz-1])+" ind: "+str(ar.index(ar[iz-1])))
                ar[iz-1]=b2
                print("("+str(ar[iz-1])+") ind: "+str(ar.index(ar[iz-1])))
                # def prnt(curl):
                # print(curl)
                # print(curl)
                zz=zz+1
                # ar.append(b2)
                # ar[iz-1]=b2
                print("IND "+str(ar.index(ar[iz-1])))
                print(iz)
                print(b2)
                print(ar[iz-1])
            
            df = pd.DataFrame({'Номер': [1, 2, 3,4, 5,6,7,8,9,10,11,12,13,14,15,16,17,18,19],
                   'Сайт': ["компаниирф.рф",
                  "https://spravkainform.ru/","http://otzyvgid.ru/","http://baza-r.ru/",
                  "https://spravkarf24.ru/",
                  "https://topserver.ru/",
                  "http://medfirms.ru/",
                  "https://feech.org/","https://korden.org/",
                  "https://rydo.ru/","https://firmexpert.ru/",
                  "https://abinsk.tvoyaspravka.ru/",
                  "https://mycompany.su/",
                  "https://spravkaru.info/","https://fis.ru/",
                  "https://www.opt-union.ru/",
                  "телефонырф.рф",
                  "https://msk.damspravku.ru/","https://1000dosok.ru/"],
                   'Карточка': [ar[0], ar[1],ar[2], ar[3], ar[4], ar[5], ar[6], ar[7], ar[8], ar[9], ar[10], ar[11], ar[12], ar[13], ar[14], ar[15],ar[16], ar[17], ar[18]],
                   'Статус': ["Готово",
                  "Готово","Готово","Готово",
                  "Готово",
                  "Готово",
                  "Готово",
                  "Готово","Готово",
                  "Готово","Готово",
                  "Готово",
                  "Готово",
                  "Готово","Модерация",
                  "Модерация",
                  "Готово",
                  "Готово","Готово"]})
            df.to_excel('./express-result.xlsx')
                
def tatabl(uuu):
    print(uuu)

eel.start("index.html", size=(930,870))



